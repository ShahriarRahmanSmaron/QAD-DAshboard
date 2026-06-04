"""Export builders for operational and chart data.

The export layer intentionally calls the existing query repository instead of
reimplementing reporting rules. That keeps exported totals aligned with the
screens that managers inspect.
"""

from __future__ import annotations

from collections.abc import Iterable
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from typing import Any

from openpyxl import Workbook  # type: ignore[import-untyped]
from openpyxl.styles import Alignment, Font, PatternFill  # type: ignore[import-untyped]
from openpyxl.utils import get_column_letter  # type: ignore[import-untyped]
from sqlalchemy.ext.asyncio import AsyncSession

from app.auth.schemas import AuthUser
from app.reporting import repository
from app.reporting.models import OperationalFact, ReportType
from app.reporting.repository import OperationalFactFilters

CONTENT_TYPE_XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _stringify(value: Any) -> Any:
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, date | datetime):
        return value.isoformat()
    if isinstance(value, list):
        return ", ".join(str(item) for item in value)
    if isinstance(value, dict):
        return ", ".join(f"{key}: {val}" for key, val in value.items())
    return value


def _fact_display_value(fact: OperationalFact) -> Any:
    if fact.value_type == "number":
        return _stringify(fact.value_numeric)
    if fact.value_type == "date":
        return _stringify(fact.value_date)
    if fact.value_type == "boolean":
        return fact.value_boolean
    if fact.value_numeric is not None:
        return _stringify(fact.value_numeric)
    if fact.is_formula and fact.formula:
        return fact.formula
    return fact.value_text


def _append_rows(sheet: Any, headers: list[str], rows: Iterable[Iterable[Any]]) -> None:
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F2937")
        cell.alignment = Alignment(horizontal="center")
    for row in rows:
        sheet.append([_stringify(value) for value in row])
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column_cells in sheet.columns:
        length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(
            max(length + 2, 10),
            45,
        )


def _filter_rows(filters: OperationalFactFilters, group_by: list[str] | None) -> list[list[Any]]:
    return [
        ["uploaded_file_id", filters.uploaded_file_id],
        ["buyer", filters.buyer],
        ["unit", filters.unit],
        ["sub_unit", filters.sub_unit],
        ["department", filters.department],
        ["buyer_id", filters.buyer_id],
        ["unit_id", filters.unit_id],
        ["metric", filters.metric_key],
        ["section", filters.operational_section],
        ["report_type_id", filters.report_type_id],
        ["report_date", filters.report_date],
        ["date_from", filters.date_from],
        ["date_to", filters.date_to],
        ["value_min", filters.value_min],
        ["value_max", filters.value_max],
        ["value_type", filters.value_type],
        ["classification", filters.row_classification],
        ["include_inactive", filters.include_inactive],
        ["search", filters.search],
        ["group_by", ", ".join(group_by or [])],
    ]


async def build_operational_query_xlsx(
    session: AsyncSession,
    *,
    user: AuthUser,
    filters: OperationalFactFilters,
    group_by: list[str] | None,
) -> bytes:
    _first_page, total = await repository.list_operational_facts(
        session,
        user=user,
        page=1,
        page_size=1,
        filters=filters,
    )
    facts, _ = await repository.list_operational_facts(
        session,
        user=user,
        page=1,
        page_size=max(total, 1),
        filters=filters,
    )
    grouped_rows, overall = await repository.aggregate_operational_facts(
        session,
        user=user,
        filters=filters,
        group_by=group_by,
    )
    resolved_group_by = repository.resolve_aggregation_dimensions(group_by)

    workbook = Workbook()
    
    # 1. Summary Sheet
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    
    report_type_name = "All Report Types"
    if filters.report_type_id:
        report_type = await session.get(ReportType, filters.report_type_id)
        if report_type:
            report_type_name = report_type.name

    active_filters = []
    if filters.metric_key: active_filters.append(f"Metric: {filters.metric_key}")
    if filters.buyer: active_filters.append(f"Buyer: {filters.buyer}")
    if filters.unit: active_filters.append(f"Unit: {filters.unit}")
    if filters.sub_unit: active_filters.append(f"Sub Unit: {filters.sub_unit}")
    if filters.department: active_filters.append(f"Department: {filters.department}")
    if filters.date_from or filters.date_to: active_filters.append(f"Date: {filters.date_from or 'Start'} -> {filters.date_to or 'End'}")
    if filters.report_date: active_filters.append(f"Selected: {filters.report_date}")
    if filters.search: active_filters.append(f"Search: {filters.search}")
    applied_filters_str = ", ".join(active_filters) if active_filters else "None"

    summary_rows = [
        ["Report Type", report_type_name],
        ["Export Timestamp", datetime.now().isoformat()],
        ["User", user.full_name or user.email],
        ["Applied Filters", applied_filters_str],
        ["Detail Record Count", total],
        ["Grouped Record Count", len(grouped_rows)],
    ]
    _append_rows(summary_sheet, ["Summary Field", "Value"], summary_rows)

    # 2. Grouped Data Sheet
    grouped_sheet = workbook.create_sheet("Grouped Data")
    grouped_headers = [
        *resolved_group_by,
        "numeric_total",
        "fact_count",
        "numeric_count",
        "formula_count",
    ]
    grouped_data = [
        [
            *(row.get(dimension) for dimension in resolved_group_by),
            row.get("numeric_total"),
            row.get("fact_count"),
            row.get("numeric_count"),
            row.get("formula_count"),
        ]
        for row in grouped_rows
    ]
    _append_rows(grouped_sheet, grouped_headers, grouped_data)
    grouped_sheet.append([])
    grouped_sheet.append(["Overall total", _stringify(overall.get("numeric_total"))])
    grouped_sheet.append(["Overall facts", overall.get("fact_count")])

    # 3. Detail Data Sheet
    detail_sheet = workbook.create_sheet("Detail Data")
    detail_headers = [
        "report_date",
        "buyer",
        "unit",
        "section",
        "metric_key",
        "metric_label",
        "classification",
        "value",
        "value_type",
        "is_formula",
        "formula",
        "sheet",
        "cell",
        "workbook_id",
    ]
    detail_rows = [
        [
            fact.report_date,
            fact.buyer,
            fact.unit,
            fact.operational_section_label,
            fact.metric_key,
            fact.metric_label,
            fact.row_classification,
            _fact_display_value(fact),
            fact.value_type,
            fact.is_formula,
            fact.formula,
            fact.source_sheet_name,
            fact.source_cell_address,
            fact.uploaded_file_id,
        ]
        for fact in facts
    ]
    _append_rows(detail_sheet, detail_headers, detail_rows)

    # 4. Applied Filters Sheet
    applied_sheet = workbook.create_sheet("Applied Filters")
    date_range_str = "None"
    if filters.date_from or filters.date_to:
        date_range_str = f"{filters.date_from or 'Start'} to {filters.date_to or 'End'}"
        
    applied_rows = [
        ["Report Type", report_type_name],
        ["Metric", filters.metric_key or "All Metrics"],
        ["Buyer", filters.buyer or "All Buyers"],
        ["Unit", filters.unit or "All Units"],
        ["Sub Unit", filters.sub_unit or "All Sub Units"],
        ["Department", filters.department or "All Departments"],
        ["Date Range", date_range_str],
        ["Selected Dates", _stringify(filters.report_date) if filters.report_date else "None"],
        ["Active Search Filters", filters.search or "None"],
    ]
    _append_rows(applied_sheet, ["Filter Name", "Value"], applied_rows)

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


async def build_trend_xlsx(
    session: AsyncSession,
    *,
    user: AuthUser,
    metric_key: str,
    buyer: str | None = None,
    unit: str | None = None,
    sub_unit: str | None = None,
    department: str | None = None,
    operational_section: str | None = None,
    report_type_id: Any = None,
    date_from: date | None = None,
    date_to: date | None = None,
    classification: str | None = None,
    series_by: str | None = None,
    limit: int = 365,
) -> bytes:
    rows = await repository.get_operational_trend(
        session,
        user=user,
        metric_key=metric_key,
        buyer=buyer,
        unit=unit,
        sub_unit=sub_unit,
        department=department,
        operational_section=operational_section,
        report_type_id=report_type_id,
        date_from=date_from,
        date_to=date_to,
        classification=classification,
        series_by=series_by,
        limit=limit,
    )

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Trend Dataset"
    headers = [
        "Date",
        "Series Name",
        "Metric Value",
        "Fact Count",
        "Numeric Count",
        "Workbooks",
    ]
    _append_rows(
        sheet,
        headers,
        (
            [
                row.get("report_date"),
                row.get("series"),
                row.get("numeric_total"),
                row.get("fact_count"),
                row.get("numeric_count"),
                row.get("workbook_names"),
            ]
            for row in rows
        ),
    )

    meta = workbook.create_sheet("Export Context")
    _append_rows(
        meta,
        ["Field", "Value"],
        [
            ["metric", metric_key],
            ["buyer", buyer],
            ["unit", unit],
            ["sub_unit", sub_unit],
            ["department", department],
            ["section", operational_section],
            ["report_type_id", report_type_id],
            ["date_from", date_from],
            ["date_to", date_to],
            ["classification", classification],
            ["series_by", series_by],
            ["limit", limit],
        ],
    )

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
