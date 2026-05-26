from __future__ import annotations

import hashlib
import json
from typing import Any

from openpyxl.utils.cell import get_column_letter

EDITABLE_REGION_KINDS = {"operational_block", "metric_zone"}
READONLY_REGION_KINDS = {
    "readonly_band",
    "summary_band",
    "formula_row",
    "calculated_row",
    "footer_region",
}
STRUCTURAL_REGION_KINDS = {
    "grouped_section",
    "section_header",
    "worksheet_separator",
    "merged_cell_region",
}


def _region_role(kind: str) -> str:
    if kind in EDITABLE_REGION_KINDS:
        return "editable"
    if kind in READONLY_REGION_KINDS:
        return "readonly"
    return "structural"


def _layout_fingerprint(value: dict[str, Any]) -> str:
    encoded = json.dumps(value, sort_keys=True, separators=(",", ":"), default=str)
    return hashlib.sha256(encoded.encode("utf-8")).hexdigest()[:24]


def _cell_address(row: int, column: int) -> str:
    return f"{get_column_letter(column)}{row}"


def _region_ids_for_point(
    regions: list[dict[str, Any]],
    *,
    row: int,
    column: int | None = None,
    include_merged: bool = True,
) -> list[str]:
    region_ids: list[str] = []
    for region in regions:
        if not include_merged and region["kind"] == "merged_cell_region":
            continue
        if row < region["start_row"] or row > region["end_row"]:
            continue
        if column is not None and (
            column < region["start_column"] or column > region["end_column"]
        ):
            continue
        region_ids.append(region["id"])
    return region_ids


def _role_for_regions(regions: list[dict[str, Any]], region_ids: list[str]) -> str:
    region_by_id = {region["id"]: region for region in regions}
    roles = {
        _region_role(region_by_id[region_id]["kind"])
        for region_id in region_ids
        if region_id in region_by_id
    }
    if "structural" in roles:
        return "structural"
    if "readonly" in roles:
        return "readonly"
    if "editable" in roles:
        return "editable"
    return "structural"


def _merged_lookup(
    sheet: Any,
) -> tuple[dict[tuple[int, int], dict[str, Any]], list[dict[str, Any]]]:
    cells: dict[tuple[int, int], dict[str, Any]] = {}
    regions: list[dict[str, Any]] = []
    for merged_range in sheet.merged_cells.ranges:
        master_address = _cell_address(merged_range.min_row, merged_range.min_col)
        span = {
            "rows": merged_range.max_row - merged_range.min_row + 1,
            "columns": merged_range.max_col - merged_range.min_col + 1,
        }
        regions.append(
            {
                "range": merged_range.coord,
                "master": master_address,
                "start_row": merged_range.min_row,
                "end_row": merged_range.max_row,
                "start_column": merged_range.min_col,
                "end_column": merged_range.max_col,
                "span": span,
            }
        )

        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for column in range(merged_range.min_col, merged_range.max_col + 1):
                role = (
                    "master"
                    if row == merged_range.min_row and column == merged_range.min_col
                    else "covered"
                )
                cells[(row, column)] = {
                    "role": role,
                    "master": master_address,
                    "range": merged_range.coord,
                    "span": span if role == "master" else {"rows": 1, "columns": 1},
                }
    return cells, regions


def build_sheet_sync_map(
    sheet: Any,
    *,
    cells: list[dict[str, Any]],
    regions: list[dict[str, Any]],
    structure: dict[str, Any],
    freeze_panes: dict[str, Any],
    preview_row_limit: int,
    preview_column_limit: int,
    orphan_masters: list[dict[str, Any]] | None = None,
) -> dict[str, Any]:
    hidden_rows = set(structure["hidden_rows"])
    hidden_columns = set(structure["hidden_columns"])
    row_heights = structure["row_heights"]
    column_widths = structure["column_widths"]
    row_groups = structure["row_groups"]
    column_groups = structure["column_groups"]
    frozen_columns = int(freeze_panes.get("frozen_columns") or 0)
    cell_lookup = {(cell["row"], cell["column"]): cell for cell in cells}
    merge_by_cell, merged_regions = _merged_lookup(sheet)
    region_by_id = {region["id"]: region for region in regions}

    # Build a fast lookup of orphan masters keyed by their first visible
    # covered cell. The frontend uses this to "rescue" the master text from a
    # hidden anchor row into the first visible covered cell of the merge so
    # operational header titles do not vanish purely because their anchor
    # row is hidden by Excel.
    orphan_masters = orphan_masters or []
    orphan_first_visible: dict[tuple[int, int], dict[str, Any]] = {}
    orphan_master_lookup: dict[str, dict[str, Any]] = {}
    for orphan in orphan_masters:
        try:
            row = int(orphan["first_visible_row"])
            column = int(orphan["first_visible_column"])
        except (KeyError, TypeError, ValueError):
            continue
        orphan_first_visible[(row, column)] = orphan
        master_address = orphan.get("master_address")
        if isinstance(master_address, str):
            orphan_master_lookup[master_address] = orphan

    row_map: list[dict[str, Any]] = []
    for row_number in range(1, preview_row_limit + 1):
        region_ids = _region_ids_for_point(
            regions,
            row=row_number,
            include_merged=False,
        )
        role = _role_for_regions(regions, region_ids)
        row_map.append(
            {
                "workbook_row": row_number,
                "grid_row_id": f"{sheet.title}::r{row_number}",
                "role": role,
                "editable": role == "editable" and row_number not in hidden_rows,
                "hidden": row_number in hidden_rows,
                "height": row_heights.get(str(row_number)),
                "outline_level": row_groups.get(str(row_number), 0),
                "region_ids": region_ids,
            }
        )

    column_map: list[dict[str, Any]] = []
    for column_number in range(1, preview_column_limit + 1):
        letter = get_column_letter(column_number)
        column_map.append(
            {
                "workbook_column": column_number,
                "workbook_column_name": letter,
                "grid_field": f"c{column_number}",
                "width": column_widths.get(letter),
                "hidden": letter in hidden_columns,
                "outline_level": column_groups.get(letter, 0),
                "frozen": column_number <= frozen_columns,
            }
        )

    editable_regions: list[dict[str, Any]] = []
    readonly_regions: list[dict[str, Any]] = []
    structural_regions: list[dict[str, Any]] = []
    for region in regions:
        target = {
            "editable": editable_regions,
            "readonly": readonly_regions,
            "structural": structural_regions,
        }[_region_role(region["kind"])]
        target.append(
            {
                "id": region["id"],
                "kind": region["kind"],
                "range": region["range"],
                "start_row": region["start_row"],
                "end_row": region["end_row"],
                "start_column": region["start_column"],
                "end_column": region["end_column"],
            }
        )

    cell_map: list[dict[str, Any]] = []
    for row_number in range(1, preview_row_limit + 1):
        for column_number in range(1, preview_column_limit + 1):
            column_name = get_column_letter(column_number)
            address = _cell_address(row_number, column_number)
            cell = cell_lookup.get((row_number, column_number))
            region_ids = _region_ids_for_point(
                regions,
                row=row_number,
                column=column_number,
                include_merged=True,
            )
            non_merged_region_ids = [
                region_id
                for region_id in region_ids
                if region_by_id.get(region_id, {}).get("kind") != "merged_cell_region"
            ]
            row_role = _role_for_regions(regions, non_merged_region_ids)
            merge = merge_by_cell.get((row_number, column_number))
            has_formula = bool(cell and cell.get("formula"))
            readonly_reason = None

            if row_number in hidden_rows or column_name in hidden_columns:
                readonly_reason = "hidden_geometry"
            elif merge and merge["role"] == "covered":
                readonly_reason = "merged_covered_cell"
            elif has_formula:
                readonly_reason = "formula"
            elif row_role == "structural":
                readonly_reason = "structural_region"
            elif row_role == "readonly":
                readonly_reason = "readonly_region"

            editable = readonly_reason is None and row_role == "editable"

            # If this cell is the first visible covered cell of an orphan
            # merge (master row hidden), surface the orphan-master metadata
            # so the frontend can render the master's text here.
            orphan_payload: dict[str, Any] | None = None
            if (row_number, column_number) in orphan_first_visible:
                orphan = orphan_first_visible[(row_number, column_number)]
                orphan_payload = {
                    "master_address": orphan.get("master_address"),
                    "range": orphan.get("range"),
                }

            cell_map.append(
                {
                    "address": cell["address"] if cell else address,
                    "workbook_row": row_number,
                    "workbook_column": column_number,
                    "grid_row_id": f"{sheet.title}::r{row_number}",
                    "grid_field": f"c{column_number}",
                    "region_ids": region_ids,
                    "editable": editable,
                    "readonly_reason": readonly_reason,
                    "has_formula": has_formula,
                    "merge": merge,
                    "blank": cell is None or cell.get("value") is None,
                    "orphan_master": orphan_payload,
                }
            )

    layout = {
        "dimension": sheet.calculate_dimension(),
        "merged_cells": structure["merged_cells"],
        "row_heights": row_heights,
        "column_widths": column_widths,
        "default_row_height": structure.get("default_row_height"),
        "default_column_width": structure.get("default_column_width"),
        "sheet_format": structure.get("sheet_format", {}),
        "hidden_rows": structure["hidden_rows"],
        "hidden_columns": structure["hidden_columns"],
        "freeze_panes": freeze_panes,
        "row_groups": row_groups,
        "column_groups": column_groups,
        "sheet_state": structure["sheet_state"],
    }

    return {
        "version": 1,
        "grid_engine": "ag-grid",
        "sheet_name": sheet.title,
        "layout_fingerprint": _layout_fingerprint(layout),
        "preview_limits": {
            "max_rows": preview_row_limit,
            "max_columns": preview_column_limit,
        },
        "geometry": layout,
        "regions": {
            "editable": editable_regions,
            "readonly": readonly_regions,
            "structural": structural_regions,
            "merged": merged_regions,
        },
        "rows": row_map,
        "columns": column_map,
        "cells": cell_map,
        "orphan_masters": list(orphan_masters),
    }


def build_workbook_sync_summary(
    *,
    filename: str,
    sheets: list[dict[str, Any]],
    parser: str,
) -> dict[str, Any]:
    summarized: list[dict[str, Any]] = []
    for sheet in sheets:
        sync = sheet.get("sync") or {}
        regions = sync.get("regions") or {}
        summarized.append(
            {
                "name": sheet.get("name", ""),
                "index": sheet.get("index", 0),
                "layout_fingerprint": sync.get("layout_fingerprint", ""),
                "editable_regions": len(regions.get("editable") or []),
                "readonly_regions": len(regions.get("readonly") or []),
                "structural_regions": len(regions.get("structural") or []),
                "mapped_rows": len(sync.get("rows") or []),
                "mapped_columns": len(sync.get("columns") or []),
                "degraded": bool(sheet.get("degraded") or sync.get("degraded")),
            }
        )
    return {
        "version": 1,
        "source_filename": filename,
        "parser": parser,
        "mapping": "xlsx_to_ag_grid",
        "sheets": summarized,
    }


def build_empty_sheet_sync_map(
    *,
    sheet_name: str,
    structure: dict[str, Any] | None = None,
    freeze_panes: dict[str, Any] | None = None,
    preview_row_limit: int = 0,
    preview_column_limit: int = 0,
    degraded: bool = False,
    degraded_reason: str | None = None,
) -> dict[str, Any]:
    """Return a fully-shaped, empty sync map.

    Used as a fallback when sync generation fails so downstream consumers can
    still iterate ``sync.cells``, ``sync.rows``, ``sync.columns`` and
    ``sync.regions`` without runtime errors.
    """

    safe_structure = structure or {
        "merged_cells": [],
        "row_heights": {},
        "column_widths": {},
        "default_row_height": None,
        "default_column_width": None,
        "sheet_format": {},
        "hidden_rows": [],
        "hidden_columns": [],
        "freeze_panes": None,
        "row_groups": {},
        "column_groups": {},
        "sheet_state": "visible",
    }
    safe_freeze = freeze_panes or {"cell": None, "frozen_rows": 0, "frozen_columns": 0}

    layout = {
        "dimension": "A1",
        "merged_cells": list(safe_structure.get("merged_cells", [])),
        "row_heights": dict(safe_structure.get("row_heights", {})),
        "column_widths": dict(safe_structure.get("column_widths", {})),
        "default_row_height": safe_structure.get("default_row_height"),
        "default_column_width": safe_structure.get("default_column_width"),
        "sheet_format": dict(safe_structure.get("sheet_format", {})),
        "hidden_rows": list(safe_structure.get("hidden_rows", [])),
        "hidden_columns": list(safe_structure.get("hidden_columns", [])),
        "freeze_panes": safe_freeze,
        "row_groups": dict(safe_structure.get("row_groups", {})),
        "column_groups": dict(safe_structure.get("column_groups", {})),
        "sheet_state": safe_structure.get("sheet_state", "visible"),
    }

    return {
        "version": 1,
        "grid_engine": "ag-grid",
        "sheet_name": sheet_name,
        "layout_fingerprint": _layout_fingerprint(layout),
        "preview_limits": {
            "max_rows": max(0, int(preview_row_limit or 0)),
            "max_columns": max(0, int(preview_column_limit or 0)),
        },
        "geometry": layout,
        "regions": {
            "editable": [],
            "readonly": [],
            "structural": [],
            "merged": [],
        },
        "rows": [],
        "columns": [],
        "cells": [],
        "orphan_masters": [],
        "degraded": bool(degraded),
        "degraded_reason": degraded_reason,
    }
