"""Workbook export / reconstruction pipeline for MD06-5.

This module rebuilds an editable copy of an uploaded XLSX workbook by *loading
the original file* with openpyxl and patching only the cells that the user
edited inside AG Grid. Loading the source workbook (instead of synthesising a
new one) preserves:

- workbook sheets and sheet order
- merged cell regions
- row heights / column widths
- freeze panes
- hidden rows / hidden columns
- workbook outline / grouping levels
- workbook styles, fills, fonts, alignments and borders (anything openpyxl
  retains across load/save)

We deliberately *do not* run a formula engine; values that were formulas in the
original workbook stay formulas, which is correct because the edited surface is
limited to operationally-editable cells (formula cells are reported as
``readonly_reason = "formula"`` by the workbook sync layer and thus never
appear in the edits payload).
"""

from __future__ import annotations

import io
import logging
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any
from uuid import UUID

from fastapi import HTTPException, status
from sqlalchemy.ext.asyncio import AsyncSession

from app.auth.schemas import AuthUser
from app.core.config import settings
from app.reporting.models import AuditLog, UploadedFile
from app.reporting.workbook_semantics import sync_workbook_semantics_after_export

logger = logging.getLogger("app.reporting.workbook_export")

ADDRESS_RE = re.compile(r"^([A-Z]+)([1-9][0-9]*)$")


def _storage_root() -> Path:
    """Mirror of ``workbook_service._storage_root`` so the two modules stay
    in sync. Kept private here to avoid a circular dependency on the upload
    service while still pointing at the same on-disk root.
    """

    configured_path = Path(settings.uploaded_workbook_storage_dir)
    if configured_path.is_absolute():
        return configured_path
    return Path(__file__).resolve().parents[2] / configured_path


def _resolve_storage_path(uploaded_file: UploadedFile) -> Path:
    """Resolve the absolute path of an uploaded workbook on disk safely.

    The stored path is relative to the *parent* of the storage root (this is
    how :func:`workbook_service.save_and_parse_workbook_upload` writes it). We
    guard against path traversal by ensuring the resolved path lives under the
    parent of the storage root.
    """

    storage_root = _storage_root()
    base = storage_root.parent.resolve()
    candidate = (base / uploaded_file.storage_path).resolve()
    try:
        candidate.relative_to(base)
    except ValueError as exc:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="Stored workbook path is outside the configured storage root.",
        ) from exc

    if not candidate.exists() or not candidate.is_file():
        raise HTTPException(
            status_code=status.HTTP_410_GONE,
            detail="Original uploaded workbook is no longer available on disk.",
        )

    return candidate


def _coerce_value(value: Any) -> Any:
    """Coerce JSON-friendly values back into types openpyxl understands.

    The frontend posts strings/numbers/bools/None; we recover dates from
    ISO ``YYYY-MM-DD`` strings and numerics from numeric strings. Anything we
    can't safely coerce passes through unchanged.
    """

    if value is None:
        return None
    if isinstance(value, bool | int | float):
        return value
    if isinstance(value, str):
        stripped = value.strip()
        if stripped == "":
            return None
        # Preserve formulas verbatim (rare here but defensive).
        if stripped.startswith("="):
            return stripped
        # ISO date
        try:
            if len(stripped) == 10 and stripped[4] == "-" and stripped[7] == "-":
                return date.fromisoformat(stripped)
        except ValueError:
            pass
        # ISO datetime
        try:
            if "T" in stripped or " " in stripped:
                return datetime.fromisoformat(stripped)
        except ValueError:
            pass
        # Numeric (avoid coercing obvious labels by requiring a digit start)
        if stripped[0] in "-0123456789.":
            try:
                if "." in stripped or "e" in stripped or "E" in stripped:
                    return float(stripped)
                return int(stripped)
            except ValueError:
                pass
        return value
    return value


def _validate_address(address: str) -> tuple[str, int]:
    if not isinstance(address, str):
        raise ValueError(f"Cell address must be a string, got {type(address).__name__}.")
    match = ADDRESS_RE.match(address.strip().upper())
    if not match:
        raise ValueError(f"Invalid cell address: {address!r}.")
    column_letters, row_number = match.group(1), int(match.group(2))
    return column_letters, row_number


def _safe_export_filename(original: str) -> str:
    """Suggest a download filename derived from the original workbook name."""

    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "-", (original or "workbook").strip()).strip(".-")
    if not cleaned:
        cleaned = "workbook"
    base, _, ext = cleaned.rpartition(".")
    if not ext or ext.lower() != "xlsx":
        return f"{cleaned}-edited.xlsx"
    return f"{base}-edited.xlsx"


def _editable_address_lookup(metadata: dict[str, Any] | None) -> dict[str, set[str]] | None:
    """Build a sheet -> editable addresses map from stored workbook sync metadata."""

    if not isinstance(metadata, dict):
        return None

    sheets = metadata.get("sheets")
    if not isinstance(sheets, list):
        return None

    lookup: dict[str, set[str]] = {}
    found_sync = False
    for sheet in sheets:
        if not isinstance(sheet, dict):
            continue
        sheet_name = sheet.get("name")
        sync = sheet.get("sync")
        if not isinstance(sheet_name, str) or not isinstance(sync, dict):
            continue
        cells = sync.get("cells")
        if not isinstance(cells, list):
            continue
        found_sync = True
        editable_addresses: set[str] = set()
        for cell in cells:
            if not isinstance(cell, dict):
                continue
            address = cell.get("address")
            if isinstance(address, str) and cell.get("editable") is True:
                editable_addresses.add(address.upper())
        lookup[sheet_name] = editable_addresses

    return lookup if found_sync else None


def _apply_sheet_edits(
    workbook: Any,
    *,
    sheet_name: str,
    edits: dict[str, Any],
    editable_addresses: set[str] | None,
    summary: dict[str, Any],
) -> None:
    if sheet_name not in workbook.sheetnames:
        summary.setdefault("missing_sheets", []).append(sheet_name)
        return

    sheet = workbook[sheet_name]
    merged_master_cells: dict[tuple[int, int], str] = {}
    covered_cells: set[tuple[int, int]] = set()
    for merged_range in sheet.merged_cells.ranges:
        merged_master_cells[(merged_range.min_row, merged_range.min_col)] = merged_range.coord
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for column in range(merged_range.min_col, merged_range.max_col + 1):
                if (row, column) == (merged_range.min_row, merged_range.min_col):
                    continue
                covered_cells.add((row, column))

    sheet_summary = summary.setdefault("sheets", {}).setdefault(
        sheet_name, {"applied": 0, "skipped": [], "patched_cells": []}
    )

    for raw_address, raw_value in edits.items():
        try:
            column_letters, row_number = _validate_address(raw_address)
        except ValueError as exc:
            sheet_summary["skipped"].append({"address": raw_address, "reason": str(exc)})
            continue
        normalized_address = f"{column_letters}{row_number}"

        if editable_addresses is not None and normalized_address not in editable_addresses:
            sheet_summary["skipped"].append(
                {"address": raw_address, "reason": "not_editable_by_sync"}
            )
            continue

        try:
            cell = sheet[normalized_address]
        except Exception as exc:  # noqa: BLE001 - openpyxl can raise various errors
            sheet_summary["skipped"].append(
                {"address": raw_address, "reason": f"cell_lookup_error:{type(exc).__name__}"}
            )
            continue

        if (cell.row, cell.column) in covered_cells:
            sheet_summary["skipped"].append(
                {"address": raw_address, "reason": "merged_covered_cell"}
            )
            continue

        existing_value = cell.value
        if isinstance(existing_value, str) and existing_value.startswith("="):
            sheet_summary["skipped"].append({"address": raw_address, "reason": "formula"})
            continue

        coerced = _coerce_value(raw_value)
        cell.value = coerced
        sheet_summary["applied"] += 1
        sheet_summary["patched_cells"].append(normalized_address)


def export_uploaded_workbook(
    *,
    uploaded_file: UploadedFile,
    edits: dict[str, dict[str, Any]] | None,
) -> tuple[bytes, str, dict[str, Any]]:
    """Reconstruct an XLSX workbook by patching the original with edits.

    Returns the binary XLSX content, a suggested download filename, and a
    structured summary describing what was applied / skipped (for audit logs
    and frontend feedback).
    """

    try:
        from openpyxl import load_workbook  # type: ignore[import-untyped]
    except ImportError as exc:  # pragma: no cover - dependency guard
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail="openpyxl is required for workbook export.",
        ) from exc

    storage_path = _resolve_storage_path(uploaded_file)
    summary: dict[str, Any] = {
        "source_filename": uploaded_file.original_filename,
        "uploaded_file_id": str(uploaded_file.id),
        "sheets": {},
        "missing_sheets": [],
        "total_edits_received": 0,
    }
    editable_lookup = _editable_address_lookup(uploaded_file.metadata_)
    summary["sync_validated"] = editable_lookup is not None

    try:
        workbook = load_workbook(storage_path, data_only=False, keep_vba=False)
    except Exception as exc:  # noqa: BLE001 - openpyxl can raise many error types
        logger.warning(
            "workbook export load failed: workbook=%r path=%s error=%s",
            uploaded_file.original_filename,
            storage_path,
            exc,
            exc_info=True,
        )
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="Unable to reopen the uploaded workbook for export.",
        ) from exc

    if edits:
        for sheet_name, sheet_edits in edits.items():
            if not isinstance(sheet_edits, dict):
                summary["sheets"].setdefault(
                    sheet_name, {"applied": 0, "skipped": [], "patched_cells": []}
                )["skipped"].append(
                    {"address": "*", "reason": "sheet_payload_not_object"}
                )
                continue
            summary["total_edits_received"] += len(sheet_edits)
            _apply_sheet_edits(
                workbook,
                sheet_name=sheet_name,
                edits=sheet_edits,
                editable_addresses=(
                    editable_lookup.get(sheet_name) if editable_lookup is not None else None
                ),
                summary=summary,
            )

    buffer = io.BytesIO()
    try:
        workbook.save(buffer)
    except Exception as exc:  # noqa: BLE001
        logger.exception(
            "workbook export save failed: workbook=%r",
            uploaded_file.original_filename,
        )
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Unable to write the reconstructed workbook.",
        ) from exc
    finally:
        workbook.close()

    payload = buffer.getvalue()
    summary["bytes_written"] = len(payload)
    summary["applied_total"] = sum(
        sheet.get("applied", 0) for sheet in summary["sheets"].values()
    )
    summary["skipped_total"] = sum(
        len(sheet.get("skipped", []) or []) for sheet in summary["sheets"].values()
    )

    download_filename = _safe_export_filename(uploaded_file.original_filename)
    return payload, download_filename, summary


async def export_workbook_for_user(
    session: AsyncSession,
    *,
    uploaded_file_id: UUID,
    edits: dict[str, dict[str, Any]] | None,
    actor: AuthUser,
) -> tuple[bytes, str, dict[str, Any]]:
    """Service entrypoint for the export endpoint.

    Looks up the uploaded file, ensures access, runs the export, and emits a
    workbook export audit log entry.
    """

    uploaded_file = await session.get(UploadedFile, uploaded_file_id)
    if uploaded_file is None or uploaded_file.deleted_at is not None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Uploaded workbook not found.",
        )

    # Owners and admins can export. Editors who uploaded the workbook see it
    # in the upload panel and should be able to roundtrip it.
    if uploaded_file.uploaded_by_user_id and uploaded_file.uploaded_by_user_id != actor.id:
        from app.auth.constants import UserRole

        if actor.role != UserRole.ADMIN:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="You can only export workbooks you uploaded.",
            )

    payload, filename, summary = export_uploaded_workbook(
        uploaded_file=uploaded_file,
        edits=edits,
    )
    semantic_sync = await sync_workbook_semantics_after_export(
        session,
        uploaded_file=uploaded_file,
        actor=actor,
        edits=edits,
        export_summary=summary,
    )

    session.add(
        AuditLog(
            actor_id=actor.id,
            actor_user_id=actor.id,
            action="workbook.exported",
            entity_type="uploaded_file",
            entity_id=str(uploaded_file.id),
            target_type="uploaded_file",
            target_id=uploaded_file.id,
            metadata_={
                "original_filename": uploaded_file.original_filename,
                "download_filename": filename,
                "applied_total": summary["applied_total"],
                "skipped_total": summary["skipped_total"],
                "bytes_written": summary["bytes_written"],
                "edited_sheets": sorted(
                    sheet
                    for sheet, info in summary["sheets"].items()
                    if info.get("applied", 0) > 0
                ),
                "missing_sheets": summary.get("missing_sheets", []),
                "semantic_fact_count": (
                    len(semantic_sync.facts) if semantic_sync is not None else None
                ),
            },
        )
    )

    return payload, filename, summary
