from __future__ import annotations

import logging
import os
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

from app.reporting.workbook_sync import (
    build_empty_sheet_sync_map,
    build_sheet_sync_map,
    build_workbook_sync_summary,
)

# ``get_column_letter`` is used to derive merged-master addresses for
# diagnostics. Imported lazily-friendly at module top because openpyxl is a
# runtime dependency of the parser anyway.
try:  # pragma: no cover - import guard exercised only when openpyxl is missing
    from openpyxl.utils.cell import get_column_letter
except ImportError:  # pragma: no cover - resolved at parse time
    def get_column_letter(index: int) -> str:
        # Minimal fallback that mirrors openpyxl's algorithm so diagnostics
        # never crash even if openpyxl fails to import (the actual parser
        # entrypoint will raise a clearer error in that case).
        if index < 1:
            return ""
        result = ""
        value = index
        while value > 0:
            value, remainder = divmod(value - 1, 26)
            result = chr(65 + remainder) + result
        return result

logger = logging.getLogger("app.reporting.workbook")

# Lightweight per-process toggle for verbose reconstruction diagnostics. The
# feature flag is intentionally cheap (env-driven) so operators can flip it on
# in production without redeploying. ``WORKBOOK_DEBUG_RECONSTRUCTION`` enables
# additional ``logger.debug`` traces for skipped rows, hidden geometry, orphan
# masters, and filtered regions. The structured warnings returned in
# ``reconstruction_diagnostics`` are emitted regardless of this flag so the
# frontend always has machine-readable visibility.
WORKBOOK_DEBUG_RECONSTRUCTION = os.environ.get(
    "WORKBOOK_DEBUG_RECONSTRUCTION", ""
).strip().lower() in {"1", "true", "yes", "on"}

MAX_PREVIEW_ROWS = 160
MAX_PREVIEW_COLUMNS = 60
MAX_REGION_ROWS_GAP = 1
HEADER_TEXT_MARKERS = ("report date", "buyer-wise", "summary", "stock", "test", "shade")
SUMMARY_TEXT_MARKERS = ("total", "grand total", "previous day", "running day")
FOOTER_TEXT_MARKERS = ("prepared by", "checked by", "approved by", "signature", "remarks", "note")


def _cell_value(value: Any) -> str | int | float | bool | None:
    if value is None:
        return None
    if isinstance(value, bool | int | float | str):
        return value
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, datetime | date):
        return value.isoformat()
    return str(value)


def _color_value(color: Any) -> str | None:
    if color is None:
        return None
    if color.type == "rgb" and color.rgb:
        return color.rgb
    if color.type == "indexed":
        return f"indexed:{color.indexed}"
    if color.type == "theme":
        return f"theme:{color.theme}"
    return None


def _border_side(side: Any) -> dict[str, Any] | None:
    if side is None or side.style is None:
        return None
    return {
        "style": side.style,
        "color": _color_value(side.color),
    }


def _has_visual_style(cell: Any) -> bool:
    if getattr(cell, "has_style", False):
        return True
    fill = getattr(cell, "fill", None)
    if fill is not None and getattr(fill, "fill_type", None):
        return True
    border = getattr(cell, "border", None)
    if border is not None and any(
        getattr(getattr(border, side, None), "style", None)
        for side in ("left", "right", "top", "bottom")
    ):
        return True
    font = getattr(cell, "font", None)
    if font is not None and (getattr(font, "bold", False) or getattr(font, "italic", False)):
        return True
    alignment = getattr(cell, "alignment", None)
    return bool(
        alignment is not None
        and (
            getattr(alignment, "horizontal", None)
            or getattr(alignment, "vertical", None)
            or getattr(alignment, "wrap_text", False)
        )
    )


def _is_formula_value(value: Any) -> bool:
    return isinstance(value, str) and value.startswith("=")


def _cell_style(cell: Any) -> dict[str, Any]:
    fill = cell.fill
    font = cell.font
    alignment = cell.alignment
    border = cell.border
    return {
        "style_id": cell.style_id,
        "number_format": cell.number_format,
        "fill": {
            "type": fill.fill_type,
            "fg_color": _color_value(fill.fgColor),
        },
        "font": {
            "bold": bool(font.bold),
            "italic": bool(font.italic),
            "color": _color_value(font.color),
            "name": font.name,
            "size": font.sz,
        },
        "alignment": {
            "horizontal": alignment.horizontal,
            "vertical": alignment.vertical,
            "wrap_text": bool(alignment.wrap_text),
        },
        "border": {
            "left": _border_side(border.left),
            "right": _border_side(border.right),
            "top": _border_side(border.top),
            "bottom": _border_side(border.bottom),
        },
    }


def _normalized_region_kind(sheet: Any, start_row: int, end_row: int) -> str:
    values: list[str] = []
    formula_count = 0
    styled_count = 0
    filled_count = 0
    bold_count = 0

    for row in sheet.iter_rows(min_row=start_row, max_row=end_row):
        for cell in row:
            if _has_visual_style(cell):
                styled_count += 1
            fill = getattr(cell, "fill", None)
            if fill is not None and getattr(fill, "fill_type", None):
                filled_count += 1
            font = getattr(cell, "font", None)
            if font is not None and getattr(font, "bold", False):
                bold_count += 1
            if cell.value is None:
                continue
            text_value = str(cell.value).strip().lower()
            values.append(text_value)
            if _is_formula_value(cell.value):
                formula_count += 1

    joined = " ".join(values)
    if not values and styled_count:
        return "worksheet_separator"
    if any(marker in joined for marker in FOOTER_TEXT_MARKERS):
        return "footer_region"
    if any(marker in joined for marker in SUMMARY_TEXT_MARKERS):
        return "summary_band"
    if formula_count > 0 and formula_count >= max(1, len(values) // 4):
        return "formula_row"
    if any(marker in joined for marker in HEADER_TEXT_MARKERS):
        return "section_header"
    if len(values) <= 3 and (filled_count or bold_count) and styled_count >= max(1, len(values)):
        return "section_header"
    if formula_count > 0:
        return "readonly_band"
    if styled_count > 0 and len(values) <= 4:
        return "metric_zone"
    return "operational_block"


def _region_metadata(sheet: Any, start_row: int, end_row: int) -> dict[str, Any]:
    formula_count = 0
    non_empty_count = 0
    style_ids: set[int] = set()
    fill_colors: set[str] = set()
    font_names: set[str] = set()
    bold_count = 0
    merged_ranges = []

    for merged_range in sheet.merged_cells.ranges:
        if merged_range.min_row <= end_row and merged_range.max_row >= start_row:
            merged_ranges.append(merged_range.coord)

    for row in sheet.iter_rows(min_row=start_row, max_row=end_row):
        for cell in row:
            if cell.value is None:
                continue
            non_empty_count += 1
            style_ids.add(cell.style_id)
            fill_color = _color_value(getattr(cell.fill, "fgColor", None))
            if fill_color:
                fill_colors.add(fill_color)
            font_name = getattr(cell.font, "name", None)
            if font_name:
                font_names.add(str(font_name))
            if getattr(cell.font, "bold", False):
                bold_count += 1
            if _is_formula_value(cell.value):
                formula_count += 1

    return {
        "formula_count": formula_count,
        "non_empty_cell_count": non_empty_count,
        "style_ids": sorted(style_ids),
        "fill_colors": sorted(fill_colors),
        "font_names": sorted(font_names),
        "bold_cell_count": bold_count,
        "merged_ranges": merged_ranges,
    }


def _build_regions(
    sheet: Any,
    *,
    diagnostics: dict[str, Any] | None = None,
) -> list[dict[str, Any]]:
    regions: list[dict[str, Any]] = []

    for index, merged_range in enumerate(sheet.merged_cells.ranges, start=1):
        regions.append(
            {
                "id": f"{sheet.title}_merged_{index}",
                "label": f"Merged {merged_range.coord}",
                "kind": "merged_cell_region",
                "range": merged_range.coord,
                "start_row": merged_range.min_row,
                "end_row": merged_range.max_row,
                "start_column": merged_range.min_col,
                "end_column": merged_range.max_col,
                "metadata": {
                    "size": {
                        "rows": merged_range.max_row - merged_range.min_row + 1,
                        "columns": merged_range.max_col - merged_range.min_col + 1,
                    }
                },
            }
        )

    # Pre-compute every row that belongs to ANY merged range so a covered row
    # (which has no own values) still anchors as an operational header during
    # band detection. Without this, merged operational headers are dropped
    # from ``active_rows`` and surrounding bands collapse incorrectly.
    merged_row_set: set[int] = set()
    for merged_range in sheet.merged_cells.ranges:
        for row_number in range(merged_range.min_row, merged_range.max_row + 1):
            merged_row_set.add(row_number)

    active_rows: list[int] = []
    skipped_blank_rows: list[int] = []
    for row in sheet.iter_rows():
        row_number = row[0].row
        has_visible_signal = any(
            cell.value is not None or _has_visual_style(cell) for cell in row
        )
        if has_visible_signal or row_number in merged_row_set:
            active_rows.append(row_number)
        else:
            skipped_blank_rows.append(row_number)

    if diagnostics is not None:
        diagnostics["skipped_blank_rows"] = skipped_blank_rows
        diagnostics["merged_row_count"] = len(merged_row_set)

    if not active_rows:
        if diagnostics is not None:
            diagnostics["bands_built"] = 0
        return regions

    def append_band(index: int, start: int, end: int, kind: str) -> None:
        regions.append(
            {
                "id": f"{sheet.title}_band_{index}",
                "label": f"Workbook band rows {start}-{end}",
                "kind": kind,
                "range": f"{start}:{end}",
                "start_row": start,
                "end_row": end,
                "start_column": 1,
                "end_column": sheet.max_column,
                "metadata": _region_metadata(sheet, start, end),
            }
        )

    start_row = active_rows[0]
    previous_row = active_rows[0]
    current_kind = _normalized_region_kind(sheet, start_row, start_row)
    region_index = 1
    for row_number in active_rows[1:]:
        row_kind = _normalized_region_kind(sheet, row_number, row_number)
        has_gap = row_number - previous_row > MAX_REGION_ROWS_GAP + 1
        if has_gap or row_kind != current_kind:
            append_band(region_index, start_row, previous_row, current_kind)
            region_index += 1
            start_row = row_number
            current_kind = row_kind
        previous_row = row_number

    append_band(region_index, start_row, previous_row, current_kind)
    if diagnostics is not None:
        diagnostics["bands_built"] = region_index
    return regions


def _freeze_pane_metadata(sheet: Any) -> dict[str, int | str | None]:
    coordinate = sheet.freeze_panes
    if coordinate is None:
        return {"cell": None, "frozen_rows": 0, "frozen_columns": 0}

    try:
        from openpyxl.utils.cell import coordinate_to_tuple

        row, column = coordinate_to_tuple(str(coordinate))
    except Exception:
        return {"cell": str(coordinate), "frozen_rows": 0, "frozen_columns": 0}

    return {
        "cell": str(coordinate),
        "frozen_rows": max(0, row - 1),
        "frozen_columns": max(0, column - 1),
    }


def _empty_structure(sheet: Any | None = None) -> dict[str, Any]:
    sheet_state = getattr(sheet, "sheet_state", None) if sheet is not None else None
    sheet_format = getattr(sheet, "sheet_format", None) if sheet is not None else None
    default_row_height = (
        getattr(sheet_format, "defaultRowHeight", None) if sheet_format is not None else None
    )
    default_column_width = (
        getattr(sheet_format, "defaultColWidth", None) if sheet_format is not None else None
    )
    return {
        "merged_cells": [],
        "row_heights": {},
        "column_widths": {},
        "default_row_height": default_row_height,
        "default_column_width": default_column_width,
        "sheet_format": {
            "base_column_width": (
                getattr(sheet_format, "baseColWidth", None) if sheet_format is not None else None
            ),
            "default_row_height": default_row_height,
            "default_column_width": default_column_width,
        },
        "hidden_rows": [],
        "hidden_columns": [],
        "freeze_panes": None,
        "row_groups": {},
        "column_groups": {},
        "sheet_state": sheet_state or "visible",
    }


def _empty_freeze_panes() -> dict[str, Any]:
    return {"cell": None, "frozen_rows": 0, "frozen_columns": 0}


def _degraded_sheet(
    *,
    sheet_index: int,
    sheet_name: str,
    sheet: Any | None,
    filename: str,
    phase: str,
    error: BaseException,
) -> dict[str, Any]:
    """Return a minimally-populated sheet entry that won't crash downstream consumers.

    Every sheet must always expose ``sync.cells``, ``sync.rows``, ``sync.columns`` and
    ``sync.regions`` so the frontend can iterate them without runtime errors.
    """

    logger.warning(
        "workbook sync degraded: sheet=%r workbook=%r phase=%s error=%s",
        sheet_name,
        filename,
        phase,
        error,
        exc_info=True,
    )

    structure = _empty_structure(sheet)
    sync = build_empty_sheet_sync_map(
        sheet_name=sheet_name,
        structure=structure,
        freeze_panes=_empty_freeze_panes(),
        preview_row_limit=0,
        preview_column_limit=0,
        degraded=True,
        degraded_reason=f"{phase}:{type(error).__name__}",
    )

    return {
        "name": sheet_name,
        "index": sheet_index,
        "dimension": "A1",
        "max_row": 0,
        "max_column": 0,
        "non_empty_cell_count": 0,
        "formula_count": 0,
        "structure": structure,
        "regions": [],
        "cells": [],
        "workbook_view": {
            "freeze_panes": _empty_freeze_panes(),
            "grid_lines": True,
            "zoom_scale": None,
        },
        "sync": sync,
        "degraded": True,
        "degraded_reason": f"{phase}:{type(error).__name__}",
        "reconstruction_diagnostics": {
            "sheet": sheet_name,
            "warnings": [
                {
                    "code": "sheet_degraded",
                    "message": (
                        f"Sheet failed to parse during {phase}; preview is unavailable."
                    ),
                    "error": f"{type(error).__name__}: {error}",
                }
            ],
            "skipped_blank_rows": [],
            "skipped_oversized_rows": 0,
            "skipped_oversized_columns": 0,
            "hidden_row_count": 0,
            "hidden_column_count": 0,
            "merged_region_count": 0,
            "merged_master_count": 0,
            "orphan_merged_masters": [],
            "merged_rows_outside_preview": 0,
            "filtered_regions": 0,
            "bands_built": 0,
            "merged_row_count": 0,
        },
    }


def _parse_single_sheet(
    sheet: Any,
    *,
    sheet_index: int,
    filename: str,
) -> dict[str, Any]:
    """Parse a single worksheet. Always returns a dict with a fully populated ``sync``.

    On any unexpected failure, returns a degraded sheet shape so callers can keep
    iterating other sheets without losing the entire workbook upload.
    """

    sheet_name = getattr(sheet, "title", None) or f"Sheet{sheet_index}"

    # Per-sheet reconstruction diagnostics. These are intentionally machine-
    # readable so the frontend can surface them in the diagnostics panel
    # without the user having to inspect server logs. They are also emitted
    # via ``logger.warning`` so SREs can see them in production traces.
    diagnostics: dict[str, Any] = {
        "sheet": sheet_name,
        "warnings": [],
        "skipped_blank_rows": [],
        "skipped_oversized_rows": 0,
        "skipped_oversized_columns": 0,
        "hidden_row_count": 0,
        "hidden_column_count": 0,
        "merged_region_count": 0,
        "merged_master_count": 0,
        "orphan_merged_masters": [],
        "merged_rows_outside_preview": 0,
        "filtered_regions": 0,
        "bands_built": 0,
        "merged_row_count": 0,
    }

    try:
        max_row = max(int(getattr(sheet, "max_row", 0) or 0), 0)
        max_column = max(int(getattr(sheet, "max_column", 0) or 0), 0)
        preview_row_limit = min(max_row, MAX_PREVIEW_ROWS)
        preview_column_limit = min(max_column, MAX_PREVIEW_COLUMNS)

        non_empty_cell_count = 0
        formula_count = 0
        cells: list[dict[str, Any]] = []
        preview_merged_cells: set[tuple[int, int]] = set()
        merged_master_addresses: list[str] = []
        merged_rows_outside_preview = 0

        try:
            for merged_range in sheet.merged_cells.ranges:
                diagnostics["merged_region_count"] += 1
                merged_master_addresses.append(
                    f"{get_column_letter(merged_range.min_col)}{merged_range.min_row}"
                )
                if (
                    merged_range.min_row > preview_row_limit
                    or merged_range.min_col > preview_column_limit
                ):
                    merged_rows_outside_preview += 1
                    continue
                for row_number in range(
                    merged_range.min_row,
                    min(merged_range.max_row, preview_row_limit) + 1,
                ):
                    for column_number in range(
                        merged_range.min_col,
                        min(merged_range.max_col, preview_column_limit) + 1,
                    ):
                        preview_merged_cells.add((row_number, column_number))
        except Exception as merged_preview_error:  # noqa: BLE001
            logger.warning(
                "workbook sync merged preview fallback: sheet=%r workbook=%r "
                "phase=merged_preview error=%s",
                sheet_name,
                filename,
                merged_preview_error,
            )
            diagnostics["warnings"].append(
                {
                    "code": "merged_preview_failed",
                    "message": "Could not enumerate merged regions for preview.",
                    "error": str(merged_preview_error),
                }
            )

        diagnostics["merged_master_count"] = len(merged_master_addresses)
        diagnostics["merged_rows_outside_preview"] = merged_rows_outside_preview
        if merged_rows_outside_preview > 0:
            diagnostics["warnings"].append(
                {
                    "code": "merged_outside_preview",
                    "message": (
                        f"{merged_rows_outside_preview} merged region(s) start beyond "
                        f"the preview window ({preview_row_limit}r x "
                        f"{preview_column_limit}c) and may be visually clipped."
                    ),
                }
            )

        if max_row > 0 and max_column > 0:
            try:
                for row in sheet.iter_rows(max_row=max_row, max_col=max_column):
                    for cell in row:
                        has_value = cell.value is not None
                        if has_value:
                            non_empty_cell_count += 1
                            if _is_formula_value(cell.value):
                                formula_count += 1

                        inside_preview = (
                            cell.row <= preview_row_limit
                            and cell.column <= preview_column_limit
                        )
                        has_preview_geometry = (cell.row, cell.column) in preview_merged_cells
                        if not inside_preview or (
                            not has_value
                            and not _has_visual_style(cell)
                            and not has_preview_geometry
                        ):
                            continue

                        formula = cell.value if _is_formula_value(cell.value) else None
                        try:
                            style = _cell_style(cell)
                        except Exception as style_error:  # noqa: BLE001
                            logger.warning(
                                "workbook sync style fallback: sheet=%r workbook=%r "
                                "phase=cell_style address=%s error=%s",
                                sheet_name,
                                filename,
                                cell.coordinate,
                                style_error,
                            )
                            style = {}
                        cells.append(
                            {
                                "address": cell.coordinate,
                                "row": cell.row,
                                "column": cell.column,
                                "value": None if formula else _cell_value(cell.value),
                                "formula": formula,
                                "data_type": cell.data_type,
                                "style": style,
                            }
                        )
            except Exception as iter_error:  # noqa: BLE001
                logger.warning(
                    "workbook sync iter_rows fallback: sheet=%r workbook=%r "
                    "phase=iter_rows error=%s",
                    sheet_name,
                    filename,
                    iter_error,
                )

        row_heights: dict[str, float] = {}
        column_widths: dict[str, float] = {}
        hidden_rows: list[int] = []
        hidden_columns: list[str] = []
        row_groups: dict[str, int] = {}
        column_groups: dict[str, int] = {}
        sheet_format = getattr(sheet, "sheet_format", None)
        default_row_height = (
            getattr(sheet_format, "defaultRowHeight", None) if sheet_format is not None else None
        )
        default_column_width = (
            getattr(sheet_format, "defaultColWidth", None) if sheet_format is not None else None
        )
        base_column_width = (
            getattr(sheet_format, "baseColWidth", None) if sheet_format is not None else None
        )
        outline_properties = getattr(getattr(sheet, "sheet_properties", None), "outlinePr", None)

        try:
            row_heights = {
                str(index): dimension.height
                for index, dimension in sheet.row_dimensions.items()
                if dimension.height is not None
            }
            column_widths = {
                key: dimension.width
                for key, dimension in sheet.column_dimensions.items()
                if dimension.width is not None
            }
            hidden_rows = [
                index
                for index, dimension in sheet.row_dimensions.items()
                if bool(dimension.hidden)
            ]
            hidden_columns = [
                key
                for key, dimension in sheet.column_dimensions.items()
                if bool(dimension.hidden)
            ]
            row_groups = {
                str(index): dimension.outlineLevel
                for index, dimension in sheet.row_dimensions.items()
                if dimension.outlineLevel
            }
            column_groups = {
                key: dimension.outlineLevel
                for key, dimension in sheet.column_dimensions.items()
                if dimension.outlineLevel
            }
        except Exception as dim_error:  # noqa: BLE001
            logger.warning(
                "workbook sync dimensions fallback: sheet=%r workbook=%r "
                "phase=dimensions error=%s",
                sheet_name,
                filename,
                dim_error,
            )
            diagnostics["warnings"].append(
                {
                    "code": "dimensions_failed",
                    "message": "Failed to read row/column dimensions; geometry may degrade.",
                    "error": str(dim_error),
                }
            )

        diagnostics["hidden_row_count"] = len(hidden_rows)
        diagnostics["hidden_column_count"] = len(hidden_columns)

        # Orphan-master detection: any merged master cell whose row is hidden
        # while at least one covered row in the same merge is visible. The
        # frontend uses this to decide whether to surface the master text in
        # the first visible covered cell so operational headers do not vanish.
        orphan_masters: list[dict[str, Any]] = []
        try:
            hidden_row_set = set(hidden_rows)
            for merged_range in sheet.merged_cells.ranges:
                if merged_range.min_row not in hidden_row_set:
                    continue
                visible_rows = [
                    row_number
                    for row_number in range(
                        merged_range.min_row, merged_range.max_row + 1
                    )
                    if row_number not in hidden_row_set
                ]
                if not visible_rows:
                    continue
                orphan_masters.append(
                    {
                        "range": merged_range.coord,
                        "master_address": (
                            f"{get_column_letter(merged_range.min_col)}{merged_range.min_row}"
                        ),
                        "first_visible_row": visible_rows[0],
                        "first_visible_column": merged_range.min_col,
                        "first_visible_address": (
                            f"{get_column_letter(merged_range.min_col)}{visible_rows[0]}"
                        ),
                        "visible_rows": visible_rows,
                    }
                )
        except Exception as orphan_error:  # noqa: BLE001
            logger.warning(
                "workbook sync orphan-master detection fallback: sheet=%r "
                "workbook=%r phase=orphan_masters error=%s",
                sheet_name,
                filename,
                orphan_error,
            )
            diagnostics["warnings"].append(
                {
                    "code": "orphan_master_detection_failed",
                    "message": "Could not derive orphan merged-master metadata.",
                    "error": str(orphan_error),
                }
            )

        diagnostics["orphan_merged_masters"] = orphan_masters
        if orphan_masters:
            logger.info(
                "workbook reconstruction: %d orphan merged master(s) detected on "
                "sheet=%r workbook=%r",
                len(orphan_masters),
                sheet_name,
                filename,
            )

        try:
            freeze_panes = _freeze_pane_metadata(sheet)
        except Exception as freeze_error:  # noqa: BLE001
            logger.warning(
                "workbook sync freeze_panes fallback: sheet=%r workbook=%r "
                "phase=freeze_panes error=%s",
                sheet_name,
                filename,
                freeze_error,
            )
            freeze_panes = _empty_freeze_panes()

        try:
            merged_cells = [
                merged_range.coord for merged_range in sheet.merged_cells.ranges
            ]
        except Exception as merged_error:  # noqa: BLE001
            logger.warning(
                "workbook sync merged_cells fallback: sheet=%r workbook=%r "
                "phase=merged_cells error=%s",
                sheet_name,
                filename,
                merged_error,
            )
            merged_cells = []

        structure = {
            "merged_cells": merged_cells,
            "row_heights": row_heights,
            "column_widths": column_widths,
            "default_row_height": default_row_height,
            "default_column_width": default_column_width,
            "sheet_format": {
                "base_column_width": base_column_width,
                "default_row_height": default_row_height,
                "default_column_width": default_column_width,
                "outline_summary_below": getattr(outline_properties, "summaryBelow", None),
                "outline_summary_right": getattr(outline_properties, "summaryRight", None),
            },
            "hidden_rows": hidden_rows,
            "hidden_columns": hidden_columns,
            "freeze_panes": (
                str(sheet.freeze_panes) if getattr(sheet, "freeze_panes", None) else None
            ),
            "row_groups": row_groups,
            "column_groups": column_groups,
            "sheet_state": getattr(sheet, "sheet_state", None) or "visible",
        }

        try:
            regions = _build_regions(sheet, diagnostics=diagnostics)
        except Exception as region_error:  # noqa: BLE001
            logger.warning(
                "workbook sync regions fallback: sheet=%r workbook=%r "
                "phase=regions error=%s",
                sheet_name,
                filename,
                region_error,
            )
            regions = []
            diagnostics["warnings"].append(
                {
                    "code": "regions_failed",
                    "message": "Workbook region detection failed; using empty regions.",
                    "error": str(region_error),
                }
            )

        try:
            sync = build_sheet_sync_map(
                sheet,
                cells=cells,
                regions=regions,
                structure=structure,
                freeze_panes=freeze_panes,
                preview_row_limit=preview_row_limit,
                preview_column_limit=preview_column_limit,
                orphan_masters=orphan_masters,
            )
        except Exception as sync_error:  # noqa: BLE001
            logger.warning(
                "workbook sync build fallback: sheet=%r workbook=%r "
                "phase=build_sync_map error=%s",
                sheet_name,
                filename,
                sync_error,
                exc_info=True,
            )
            sync = build_empty_sheet_sync_map(
                sheet_name=sheet_name,
                structure=structure,
                freeze_panes=freeze_panes,
                preview_row_limit=preview_row_limit,
                preview_column_limit=preview_column_limit,
                degraded=True,
                degraded_reason=f"build_sync_map:{type(sync_error).__name__}",
            )
            diagnostics["warnings"].append(
                {
                    "code": "sync_build_failed",
                    "message": "Workbook sync map could not be built; sheet is degraded.",
                    "error": str(sync_error),
                }
            )

        try:
            dimension = sheet.calculate_dimension()
        except Exception:
            dimension = "A1"

        try:
            grid_lines = bool(sheet.sheet_view.showGridLines)
        except Exception:
            grid_lines = True
        try:
            zoom_scale = sheet.sheet_view.zoomScale
        except Exception:
            zoom_scale = None

        # Emit verbose debug traces only when the env flag is on. The
        # structured diagnostics dict is always returned and surfaced in the
        # workbook sync output so the frontend can display reconstruction
        # warnings without operators flipping a flag first.
        if WORKBOOK_DEBUG_RECONSTRUCTION:
            logger.debug(
                "workbook reconstruction diagnostics: workbook=%r sheet=%r "
                "hidden_rows=%d hidden_columns=%d merged_regions=%d "
                "orphan_masters=%d skipped_blank_rows=%d bands_built=%d",
                filename,
                sheet_name,
                diagnostics["hidden_row_count"],
                diagnostics["hidden_column_count"],
                diagnostics["merged_region_count"],
                len(diagnostics["orphan_merged_masters"]),
                len(diagnostics["skipped_blank_rows"]),
                diagnostics["bands_built"],
            )

        return {
            "name": sheet_name,
            "index": sheet_index,
            "dimension": dimension,
            "max_row": max_row,
            "max_column": max_column,
            "non_empty_cell_count": non_empty_cell_count,
            "formula_count": formula_count,
            "structure": structure,
            "regions": regions,
            "cells": cells,
            "workbook_view": {
                "freeze_panes": freeze_panes,
                "grid_lines": grid_lines,
                "zoom_scale": zoom_scale,
            },
            "sync": sync,
            "reconstruction_diagnostics": diagnostics,
        }
    except Exception as exc:  # noqa: BLE001
        return _degraded_sheet(
            sheet_index=sheet_index,
            sheet_name=sheet_name,
            sheet=sheet,
            filename=filename,
            phase="sheet_parse",
            error=exc,
        )


def parse_xlsx_workbook(path: Path, *, filename: str) -> dict[str, Any]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError(
            "openpyxl is required for workbook ingestion. Install backend dependencies first."
        ) from exc

    try:
        import pandas as pd

        pandas_sheet_names = list(pd.ExcelFile(path).sheet_names)
    except Exception:
        pandas_sheet_names = []

    workbook = load_workbook(path, data_only=False)
    sheets: list[dict[str, Any]] = []
    try:
        for sheet_index, sheet in enumerate(workbook.worksheets, start=1):
            sheets.append(
                _parse_single_sheet(sheet, sheet_index=sheet_index, filename=filename)
            )
    finally:
        workbook.close()

    parser_name = "openpyxl+pandas" if pandas_sheet_names else "openpyxl"
    degraded_sheets = [sheet["name"] for sheet in sheets if sheet.get("degraded")]
    if degraded_sheets:
        logger.warning(
            "workbook upload completed with degraded sheets: workbook=%r sheets=%s",
            filename,
            degraded_sheets,
        )

    # Workbook-level reconstruction diagnostics. We aggregate per-sheet
    # diagnostics so the frontend can show a single rolled-up panel without
    # having to re-traverse every sheet's metadata.
    aggregated_warnings: list[dict[str, Any]] = []
    aggregated_orphan_masters = 0
    aggregated_hidden_rows = 0
    aggregated_hidden_columns = 0
    aggregated_merged_regions = 0
    aggregated_skipped_rows = 0
    aggregated_bands_built = 0
    for sheet in sheets:
        sheet_diag = sheet.get("reconstruction_diagnostics") or {}
        for warning in sheet_diag.get("warnings", []) or []:
            aggregated_warnings.append({"sheet": sheet.get("name"), **warning})
        aggregated_orphan_masters += len(sheet_diag.get("orphan_merged_masters", []) or [])
        aggregated_hidden_rows += int(sheet_diag.get("hidden_row_count", 0) or 0)
        aggregated_hidden_columns += int(sheet_diag.get("hidden_column_count", 0) or 0)
        aggregated_merged_regions += int(sheet_diag.get("merged_region_count", 0) or 0)
        aggregated_skipped_rows += len(sheet_diag.get("skipped_blank_rows", []) or [])
        aggregated_bands_built += int(sheet_diag.get("bands_built", 0) or 0)

    return {
        "filename": filename,
        "sheet_count": len(sheets),
        "parser": parser_name,
        "preview_limits": {
            "max_rows": MAX_PREVIEW_ROWS,
            "max_columns": MAX_PREVIEW_COLUMNS,
            "pandas_sheet_names": pandas_sheet_names,
        },
        "workbook_sync": build_workbook_sync_summary(
            filename=filename,
            sheets=sheets,
            parser=parser_name,
        ),
        "sheets": sheets,
        "degraded_sheets": degraded_sheets,
        "reconstruction_diagnostics": {
            "warnings": aggregated_warnings,
            "orphan_merged_masters": aggregated_orphan_masters,
            "hidden_rows": aggregated_hidden_rows,
            "hidden_columns": aggregated_hidden_columns,
            "merged_regions": aggregated_merged_regions,
            "skipped_blank_rows": aggregated_skipped_rows,
            "bands_built": aggregated_bands_built,
            "debug_logging_enabled": WORKBOOK_DEBUG_RECONSTRUCTION,
        },
    }
