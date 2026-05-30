"""Synthetic workbook-metadata builder for ownership tests.

Produces the same sheet/cell/region shape the real parser emits, so tests can
exercise the ownership engine without depending on any specific .xlsx file.
There are **no business values baked in** — every test supplies its own grid,
keeping the engine's workbook-agnostic contract honest.
"""

from __future__ import annotations

from typing import Any

from openpyxl.utils.cell import get_column_letter

JsonObject = dict[str, Any]


def _data_type(value: Any) -> str:
    if value is None:
        return "n"
    if isinstance(value, bool):
        return "b"
    if isinstance(value, (int, float)):
        return "n"
    return "s"


def build_sheet(
    *,
    name: str,
    grid: dict[tuple[int, int], Any],
    merges: list[str] | None = None,
    formulas: dict[tuple[int, int], str] | None = None,
    index: int = 1,
) -> JsonObject:
    """Build a parser-shaped sheet dict from a sparse cell grid.

    ``grid`` maps ``(row, column)`` -> value. ``merges`` are A1-style ranges
    (e.g. ``"B5:B10"``). ``formulas`` maps ``(row, column)`` -> formula string;
    those cells are emitted with ``value=None`` and ``formula=...`` exactly as
    the parser stores computed cells.
    """
    formulas = formulas or {}
    merges = merges or []
    max_row = max((r for r, _ in grid), default=0)
    max_col = max((c for _, c in grid), default=0)
    for r, c in formulas:
        max_row = max(max_row, r)
        max_col = max(max_col, c)

    cells: list[JsonObject] = []
    seen: set[tuple[int, int]] = set()
    for (row, column), value in sorted(grid.items()):
        seen.add((row, column))
        cells.append(
            {
                "address": f"{get_column_letter(column)}{row}",
                "row": row,
                "column": column,
                "value": value,
                "formula": None,
                "data_type": _data_type(value),
                "style": {},
            }
        )
    for (row, column), formula in sorted(formulas.items()):
        if (row, column) in seen:
            continue
        cells.append(
            {
                "address": f"{get_column_letter(column)}{row}",
                "row": row,
                "column": column,
                "value": None,
                "formula": formula,
                "data_type": "n",
                "style": {},
            }
        )

    regions: list[JsonObject] = []
    for idx, coord in enumerate(merges, start=1):
        start, end = coord.split(":")
        sc, sr = _split_addr(start)
        ec, er = _split_addr(end)
        regions.append(
            {
                "id": f"{name}_merged_{idx}",
                "label": f"Merged {coord}",
                "kind": "merged_cell_region",
                "range": coord,
                "start_row": sr,
                "end_row": er,
                "start_column": sc,
                "end_column": ec,
                "metadata": {},
            }
        )

    return {
        "name": name,
        "index": index,
        "dimension": f"A1:{get_column_letter(max(max_col, 1))}{max(max_row, 1)}",
        "max_row": max_row,
        "max_column": max_col,
        "non_empty_cell_count": len(cells),
        "formula_count": len(formulas),
        "structure": {"merged_cells": list(merges)},
        "regions": regions,
        "cells": cells,
        "workbook_view": {},
        "sync": {},
    }


def build_workbook(*, filename: str, sheets: list[JsonObject]) -> JsonObject:
    return {
        "filename": filename,
        "sheet_count": len(sheets),
        "parser": "synthetic",
        "preview_limits": {},
        "workbook_sync": {},
        "sheets": sheets,
        "workbook_source": {"synthetic": True},
    }


def _split_addr(addr: str) -> tuple[int, int]:
    letters = "".join(ch for ch in addr if ch.isalpha())
    digits = "".join(ch for ch in addr if ch.isdigit())
    col = 0
    for ch in letters:
        col = col * 26 + (ord(ch.upper()) - ord("A") + 1)
    return col, int(digits)
